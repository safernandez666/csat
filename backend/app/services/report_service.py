"""PDF + XLSX report generation for CSAT.

The PDF uses reportlab and falls back to fpdf2 if reportlab is missing. The
layout targets an executive deliverable: a cover with a real donut score,
visual KPI cards for IG maturity, color-coded status chips, and a per-control
breakdown grouped by Implementation Group.
"""
import os
from io import BytesIO
from datetime import datetime, timezone
from sqlalchemy.orm import Session

from app.core.config import settings as app_settings
from app.models.control import Control, ControlStatus, Safeguard
from app.models.evidence import Evidence
from app.models.settings import Setting
from app.models.user import User
from app.services.control_service import compute_compliance_score
from app.utils.cis_groups import get_cis_group
from app.utils.safeguard_status import count_implemented


# ---------------------------------------------------------------------------
# Settings + AI helpers
# ---------------------------------------------------------------------------

def _get_platform_name(db: Session) -> str:
    s = db.query(Setting).filter(Setting.key == "platform_name").first()
    return s.value if s and s.value else "CSAT"


def _get_logo_path(db: Session):
    """Resolve company_logo_url (e.g. /uploads/foo.png) to an absolute path
    inside the upload dir. Returns None if missing or invalid."""
    s = db.query(Setting).filter(Setting.key == "company_logo_url").first()
    if not s or not s.value:
        return None
    url = str(s.value)
    if not url.startswith("/uploads/"):
        return None
    filename = url[len("/uploads/"):]
    base = os.path.abspath(app_settings.upload_dir)
    target = os.path.abspath(os.path.join(base, filename))
    # path-traversal guard
    if not target.startswith(base + os.sep) and target != base:
        return None
    return target if os.path.isfile(target) else None


def _get_ai_config(db: Session) -> dict:
    s = db.query(Setting).filter(Setting.key == "ai_config").first()
    return s.value if s and s.value else {
        "provider": "ollama",
        "api_url": "http://localhost:11434",
        "api_key": "",
        "model": "llama3:latest",
    }


def _generate_ai_summary(db: Session, controls: list, summary: dict, ig_progress: dict, quick_wins: list) -> str:
    """Generate an executive summary using AI. Returns fallback text on failure."""
    config = _get_ai_config(db)
    if not config.get("api_url"):
        return _generate_fallback_summary(controls, summary, ig_progress, quick_wins)

    try:
        from app.connectors.ai_analysis import AIAnalysisConnector
        conn = AIAnalysisConnector()
        conn.configure(config)

        context = (
            f"Platform: {_get_platform_name(db)}\n"
            f"Compliance Score: {summary['compliance_score']}%\n"
            f"Total Controls: {summary['total']}\n"
            f"Implemented: {summary['implemented']}, In Progress: {summary['in_progress']}, "
            f"Not Implemented: {summary['not_implemented']}, Needs Review: {summary['needs_review']}\n"
            f"Risk Distribution: Critical={summary['by_risk']['critical']}, High={summary['by_risk']['high']}, "
            f"Medium={summary['by_risk']['medium']}, Low={summary['by_risk']['low']}\n"
            f"IG Maturity: IG1={ig_progress.get('ig1', {}).get('score', 0)}%, "
            f"IG2={ig_progress.get('ig2', {}).get('score', 0)}%, "
            f"IG3={ig_progress.get('ig3', {}).get('score', 0)}%\n\n"
            f"Top Quick Wins:\n"
        )
        for w in quick_wins[:5]:
            context += f"- CIS {w['cis_id']} {w['name']}: {w.get('why', '')}\n"

        messages = [
            {
                "role": "system",
                "content": (
                    "You are a senior cybersecurity compliance consultant writing an executive summary "
                    "for a CIS Controls v8 assessment report. Write 3-4 concise paragraphs summarizing "
                    "the current security posture, key risks, and top priorities. Be professional, direct, "
                    "and actionable. Do not use markdown formatting. Use plain prose only."
                ),
            },
            {"role": "user", "content": context},
        ]
        return conn.chat(messages).strip()
    except Exception:
        return _generate_fallback_summary(controls, summary, ig_progress, quick_wins)


def _generate_fallback_summary(controls: list, summary: dict, ig_progress: dict, quick_wins: list) -> str:
    score = summary["compliance_score"]
    not_impl = summary["not_implemented"]
    critical_high = summary["by_risk"]["critical"] + summary["by_risk"]["high"]

    paragraphs = [
        f"The organization currently has a compliance score of {score}% across {summary['total']} CIS Controls v8. "
        f"Out of these, {summary['implemented']} are fully implemented, {summary['in_progress']} are in progress, "
        f"and {not_impl} remain not implemented. {summary['needs_review']} control(s) require review.",
    ]
    if critical_high > 0:
        paragraphs.append(
            f"There are {critical_high} controls classified as critical or high risk that are not yet fully implemented. "
            f"These represent the most urgent gaps in the security posture and should be prioritized for immediate remediation."
        )
    else:
        paragraphs.append(
            "All controls are classified as medium or low risk. While this reduces immediate exposure, "
            "continued implementation progress is essential to maintain a robust security posture."
        )
    ig1_score = ig_progress.get("ig1", {}).get("score", 0)
    if ig1_score < 100:
        paragraphs.append(
            f"Implementation Group 1 (IG1) — the essential safeguards — is at {ig1_score}% completion. "
            f"Achieving full IG1 compliance should be the top priority as it forms the foundational security baseline."
        )
    if quick_wins:
        paragraphs.append(
            f"The top recommended quick win is CIS {quick_wins[0]['cis_id']} ({quick_wins[0]['name']}). "
            f"{quick_wins[0].get('next_action', 'Begin implementation planning immediately.')}."
        )
    return "\n\n".join(paragraphs)


def _compute_quick_wins(controls: list) -> list:
    """Heuristic quick wins (mirrors api/ai.py)."""
    candidates = []
    for c in controls:
        total = len(c.safeguards)
        implemented = count_implemented(c.safeguards)
        not_impl_ig1 = sum(1 for s in c.safeguards if s.ig == "ig1" and s.implementation_status == "not_implemented")
        not_impl_ig2 = sum(1 for s in c.safeguards if s.ig == "ig2" and s.implementation_status == "not_implemented")
        not_impl_ig3 = sum(1 for s in c.safeguards if s.ig == "ig3" and s.implementation_status == "not_implemented")

        risk_score = {"critical": 1, "high": 2, "medium": 3, "low": 4}.get(c.risk_level, 3)
        owner_score = 0 if c.owner_id else 1
        evidence_count = len(c.evidence) if hasattr(c, "evidence") else 0
        evidence_score = 0 if evidence_count > 0 else 1
        ig1_score = (not_impl_ig1 * -5) + (not_impl_ig2 * -2) + (not_impl_ig3 * -1)
        quick_win_score = risk_score + owner_score + evidence_score + ig1_score

        reasons = []
        if not_impl_ig1 > 0:
            reasons.append(f"Has {not_impl_ig1} IG1 safeguards pending")
        if c.risk_level in ("critical", "high"):
            reasons.append(f"{c.risk_level.capitalize()} risk level")
        if not c.owner_id:
            reasons.append("No owner assigned")
        if evidence_count == 0:
            reasons.append("No evidence collected yet")
        why = "; ".join(reasons) if reasons else "Control partially implemented"

        if not_impl_ig1 > 0:
            next_action = f"Start by implementing the {not_impl_ig1} IG1 safeguard(s) for this control"
        elif not c.owner_id:
            next_action = "Assign an owner and begin implementation planning"
        elif evidence_count == 0:
            next_action = "Document existing controls as evidence"
        else:
            next_action = "Continue implementing remaining safeguards"

        effort = "low" if not_impl_ig1 <= 2 and c.risk_level in ("low", "medium") else "medium" if not_impl_ig1 <= 4 else "high"
        impact = "high" if c.risk_level in ("critical", "high") or not_impl_ig1 >= 3 else "medium" if not_impl_ig1 >= 1 else "low"

        candidates.append({
            "cis_id": c.cis_id, "name": c.name, "group": get_cis_group(c.cis_id),
            "risk_level": c.risk_level, "status": c.status,
            "safeguards_total": total, "safeguards_implemented": implemented,
            "ig1_pending": not_impl_ig1, "ig2_pending": not_impl_ig2, "ig3_pending": not_impl_ig3,
            "evidence_count": evidence_count, "has_owner": bool(c.owner_id),
            "quick_win_score": quick_win_score, "objective": c.objective,
            "why": why, "next_action": next_action, "effort": effort, "impact": impact,
        })

    candidates.sort(key=lambda x: (-x["ig1_pending"], x["quick_win_score"]))
    return candidates[:10]


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def generate_pdf_report(db: Session) -> bytes:
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, KeepTogether, Image,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.lib.colors import HexColor
        from reportlab.graphics.shapes import Drawing, Rect, String, Wedge, Circle
    except ImportError:
        return _generate_with_fpdf(db)

    # ----- Aggregate data
    controls = db.query(Control).all()
    total = len(controls)
    implemented = sum(1 for c in controls if c.status == ControlStatus.IMPLEMENTED.value)
    in_progress = sum(1 for c in controls if c.status == ControlStatus.IN_PROGRESS.value)
    not_impl = sum(1 for c in controls if c.status == ControlStatus.NOT_IMPLEMENTED.value)
    needs_review = sum(1 for c in controls if c.status == ControlStatus.NEEDS_REVIEW.value)
    score = compute_compliance_score(controls)
    by_risk = {
        "critical": sum(1 for c in controls if c.risk_level == "critical"),
        "high":     sum(1 for c in controls if c.risk_level == "high"),
        "medium":   sum(1 for c in controls if c.risk_level == "medium"),
        "low":      sum(1 for c in controls if c.risk_level == "low"),
    }

    safeguards = db.query(Safeguard).all()
    ig_progress = {"ig1": {"total": 0, "implemented": 0},
                   "ig2": {"total": 0, "implemented": 0},
                   "ig3": {"total": 0, "implemented": 0}}
    for sg in safeguards:
        if sg.ig not in ig_progress:
            continue
        ig_progress[sg.ig]["total"] += 1
        if sg.implementation_status == "not_applicable":
            continue
        if sg.implementation_status == "implemented_all":
            ig_progress[sg.ig]["implemented"] += 1
    for k in ig_progress:
        t, i = ig_progress[k]["total"], ig_progress[k]["implemented"]
        ig_progress[k]["score"] = round((i / t) * 100) if t else 0

    quick_wins = _compute_quick_wins(controls)
    platform_name = _get_platform_name(db)
    logo_path = _get_logo_path(db)
    ai_summary = _generate_ai_summary(db, controls, {
        "total": total, "implemented": implemented, "in_progress": in_progress,
        "not_implemented": not_impl, "needs_review": needs_review,
        "compliance_score": score, "by_risk": by_risk,
    }, ig_progress, quick_wins)

    # ----- Color palette (mirror frontend tokens)
    BRAND       = HexColor("#18181B")
    BRAND_LIGHT = HexColor("#F4F4F5")
    SUCCESS     = HexColor("#16A34A")
    SUCCESS_BG  = HexColor("#DCFCE7")
    INFO        = HexColor("#2563EB")
    INFO_BG     = HexColor("#DBEAFE")
    WARNING     = HexColor("#D97706")
    WARNING_BG  = HexColor("#FEF3C7")
    DANGER      = HexColor("#DC2626")
    PURPLE      = HexColor("#9333EA")
    MUTED       = HexColor("#71717A")
    BORDER      = HexColor("#E4E4E7")
    ORANGE      = HexColor("#EA580C")

    def hx(c):
        return f"#{int(c.red * 255):02X}{int(c.green * 255):02X}{int(c.blue * 255):02X}"

    STATUS_BG = {"implemented": SUCCESS_BG, "in_progress": INFO_BG,
                 "needs_review": WARNING_BG, "not_implemented": BRAND_LIGHT}
    STATUS_FG = {"implemented": SUCCESS, "in_progress": INFO,
                 "needs_review": WARNING, "not_implemented": MUTED}
    RISK_FG = {"critical": DANGER, "high": ORANGE, "medium": WARNING, "low": SUCCESS}

    # ----- Styles
    styles = getSampleStyleSheet()
    cover_title_style = ParagraphStyle(
        'CoverTitle', parent=styles['Heading1'],
        fontSize=34, leading=40, textColor=BRAND, fontName='Helvetica-Bold',
        spaceAfter=4, alignment=1,
    )
    cover_subtitle_style = ParagraphStyle(
        'CoverSub', parent=styles['Normal'],
        fontSize=12, leading=16, textColor=MUTED, alignment=1, spaceAfter=24,
    )
    cover_meta_style = ParagraphStyle(
        'CoverMeta', parent=styles['Normal'],
        fontSize=9, leading=14, textColor=MUTED, alignment=1,
    )
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'],
        fontSize=16, leading=20, textColor=BRAND, fontName='Helvetica-Bold',
        spaceAfter=10, spaceBefore=14,
    )
    subsection_style = ParagraphStyle(
        'Subsection', parent=styles['Normal'],
        fontSize=11, leading=14, textColor=BRAND, fontName='Helvetica-Bold',
        spaceAfter=6, spaceBefore=10,
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['Normal'],
        fontSize=10, leading=14, spaceAfter=6, textColor=BRAND,
    )
    body_muted_style = ParagraphStyle(
        'BodyMuted', parent=body_style, textColor=MUTED, fontSize=9, leading=12,
    )

    # ----- Visual primitives

    def donut(score_pct, size=2.6 * inch):
        # ReportLab's annular Wedge has a rendering bug for full rings, so we
        # build the donut from a full ring (Circle in BORDER color), a wedge
        # for the score, then a white Circle on top to carve the hole.
        d = Drawing(size, size)
        cx = cy = size / 2
        outer_r = size / 2 - 4
        inner_r = outer_r * 0.68
        if score_pct >= 70:
            arc_color = SUCCESS
        elif score_pct >= 40:
            arc_color = WARNING
        else:
            arc_color = DANGER
        # Full background circle
        d.add(Circle(cx, cy, outer_r, fillColor=BORDER, strokeColor=None))
        # Progress wedge (clockwise from 12 o'clock)
        if score_pct > 0 and score_pct < 100:
            sweep = 360 * (score_pct / 100)
            d.add(Wedge(cx, cy, outer_r, 90 - sweep, 90,
                        fillColor=arc_color, strokeColor=None))
        elif score_pct >= 100:
            d.add(Circle(cx, cy, outer_r, fillColor=arc_color, strokeColor=None))
        # White center to create the hole
        d.add(Circle(cx, cy, inner_r, fillColor=colors.white, strokeColor=None))
        # Center labels
        d.add(String(cx, cy - 6, f"{score_pct}%",
                     fontName='Helvetica-Bold', fontSize=34,
                     fillColor=BRAND, textAnchor='middle'))
        d.add(String(cx, cy - 28, "compliance",
                     fontName='Helvetica', fontSize=9,
                     fillColor=MUTED, textAnchor='middle'))
        return d

    def ig_card(label, sub_label, score_pct, implemented_n, total_n, accent):
        w, h = 2.05 * inch, 1.5 * inch
        d = Drawing(w, h)
        d.add(Rect(0, 0, w, h, fillColor=BRAND_LIGHT, strokeColor=None, rx=8, ry=8))
        d.add(String(12, h - 18, label,
                     fontName='Helvetica-Bold', fontSize=11, fillColor=BRAND))
        d.add(String(12, h - 32, sub_label,
                     fontName='Helvetica', fontSize=8, fillColor=MUTED))
        d.add(String(12, h - 64, f"{score_pct}%",
                     fontName='Helvetica-Bold', fontSize=26, fillColor=accent))
        bar_w = w - 24
        d.add(Rect(12, 28, bar_w, 6, fillColor=BORDER, strokeColor=None, rx=3, ry=3))
        if score_pct > 0:
            d.add(Rect(12, 28, bar_w * (score_pct / 100), 6,
                       fillColor=accent, strokeColor=None, rx=3, ry=3))
        d.add(String(12, 12, f"{implemented_n} / {total_n} safeguards",
                     fontName='Helvetica', fontSize=8, fillColor=MUTED))
        return d

    def horizontal_bar(value, max_value, fill_color, width=3 * inch):
        d = Drawing(width, 14)
        d.add(Rect(0, 4, width, 6, fillColor=BORDER, strokeColor=None, rx=2, ry=2))
        if max_value > 0 and value > 0:
            d.add(Rect(0, 4, width * (value / max_value), 6,
                       fillColor=fill_color, strokeColor=None, rx=2, ry=2))
        return d

    # ----- Doc
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54,
    )
    story = []

    # ===== COVER =====
    story.append(Spacer(1, 0.7 * inch))
    if logo_path:
        try:
            logo = Image(logo_path, width=1.3 * inch, height=1.3 * inch, kind='proportional')
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 0.25 * inch))
        except Exception:
            pass

    story.append(Paragraph(platform_name, cover_title_style))
    story.append(Paragraph(
        "CIS Controls v8 &mdash; Compliance Assessment Report",
        cover_subtitle_style,
    ))
    story.append(Spacer(1, 0.4 * inch))

    cover_donut = donut(score, size=2.8 * inch)
    cover_donut.hAlign = 'CENTER'
    story.append(cover_donut)

    story.append(Spacer(1, 0.4 * inch))
    cover_meta = (
        f"Assessment generated {datetime.now(timezone.utc).strftime('%B %d, %Y at %H:%M UTC')}"
        f"<br/>{total} CIS Controls v8 &middot; {sum(s['total'] for s in ig_progress.values())} safeguards evaluated"
    )
    story.append(Paragraph(cover_meta, cover_meta_style))
    story.append(PageBreak())

    # ===== EXECUTIVE SUMMARY =====
    story.append(Paragraph("Executive Summary", section_style))
    summary_html = ai_summary.replace("\n\n", "<br/><br/>").replace("\n", " ")
    story.append(Paragraph(summary_html, body_style))
    story.append(Spacer(1, 0.2 * inch))

    # ===== COMPLIANCE OVERVIEW =====
    story.append(Paragraph("Compliance Overview", section_style))

    pct = lambda n: f"{round((n / total) * 100) if total else 0}%"
    metrics_rows = [
        ["", "Implemented",     str(implemented),  pct(implemented)],
        ["", "In Progress",     str(in_progress),  pct(in_progress)],
        ["", "Not Implemented", str(not_impl),     pct(not_impl)],
        ["", "Needs Review",    str(needs_review), pct(needs_review)],
    ]
    metrics_tbl = Table(metrics_rows,
                        colWidths=[0.18 * inch, 1.7 * inch, 0.6 * inch, 0.7 * inch])
    metrics_tbl.setStyle(TableStyle([
        ('FONTNAME', (1, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (1, 0), (-1, -1), 10),
        ('TEXTCOLOR', (1, 0), (1, -1), BRAND),
        ('TEXTCOLOR', (2, 0), (2, -1), BRAND),
        ('TEXTCOLOR', (3, 0), (3, -1), MUTED),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (1, 0), (-1, -2), 0.5, BORDER),
        ('BACKGROUND', (0, 0), (0, 0), SUCCESS),
        ('BACKGROUND', (0, 1), (0, 1), INFO),
        ('BACKGROUND', (0, 2), (0, 2), MUTED),
        ('BACKGROUND', (0, 3), (0, 3), WARNING),
    ]))
    overview = Table([[donut(score, size=2.2 * inch), metrics_tbl]],
                     colWidths=[2.6 * inch, 3.5 * inch])
    overview.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(overview)
    # Force a page break after the score block so the variable-length AI
    # summary can't push Risk Distribution into an awkward orphan layout.
    story.append(PageBreak())

    max_risk = max(by_risk.values()) if any(by_risk.values()) else 1
    risk_rows = []
    for level, label in [("critical", "Critical"), ("high", "High"),
                          ("medium", "Medium"), ("low", "Low")]:
        risk_rows.append([
            label,
            str(by_risk[level]),
            horizontal_bar(by_risk[level], max_risk, RISK_FG[level], width=3.2 * inch),
        ])
    risk_tbl = Table(risk_rows, colWidths=[1 * inch, 0.4 * inch, 3.4 * inch])
    risk_tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), BRAND),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(Paragraph("Risk Distribution", subsection_style))
    story.append(risk_tbl)
    story.append(Spacer(1, 0.25 * inch))

    # ===== IG MATURITY =====
    story.append(Paragraph("Implementation Group Maturity", section_style))
    story.append(Paragraph(
        "Maturity progress across CIS Implementation Groups (Waves). IG1 is the essential baseline; IG2 and IG3 represent intermediate and advanced safeguards.",
        body_muted_style,
    ))
    story.append(Spacer(1, 0.18 * inch))

    ig1, ig2, ig3 = ig_progress["ig1"], ig_progress["ig2"], ig_progress["ig3"]
    ig_row = Table([[
        ig_card("IG1", "Wave 1 — Essential",     ig1["score"], ig1["implemented"], ig1["total"], SUCCESS),
        ig_card("IG2", "Wave 2 — Intermediate",  ig2["score"], ig2["implemented"], ig2["total"], INFO),
        ig_card("IG3", "Wave 3 — Advanced",      ig3["score"], ig3["implemented"], ig3["total"], PURPLE),
    ]], colWidths=[2.15 * inch, 2.15 * inch, 2.15 * inch])
    ig_row.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(ig_row)
    # Quick Wins gets its own page so the cards have room to breathe.
    story.append(PageBreak())

    # ===== QUICK WINS =====
    if quick_wins:
        story.append(Paragraph("Top Quick Wins", section_style))
        story.append(Paragraph(
            "Prioritized recommendations based on risk level, IG1 pending safeguards, and implementation effort. Addressing these first yields the highest security improvement per unit of effort.",
            body_muted_style,
        ))
        story.append(Spacer(1, 0.15 * inch))

        badges_style = ParagraphStyle(
            'Badges', parent=body_muted_style, fontSize=8, leading=10,
            spaceAfter=0, spaceBefore=0,
        )
        for i, w in enumerate(quick_wins[:5], 1):
            effort_color = SUCCESS if w["effort"] == "low" else WARNING if w["effort"] == "medium" else DANGER
            impact_color = SUCCESS if w["impact"] == "high" else WARNING if w["impact"] == "medium" else DANGER
            risk_color = RISK_FG.get(w["risk_level"], MUTED)

            header_html = (
                f'<font name="Helvetica-Bold" size="11" color="{hx(BRAND)}">'
                f'#{i} &nbsp; CIS {w["cis_id"]} &middot; {w["name"]}</font>'
            )
            badges_html = (
                f'<font color="{hx(risk_color)}"><b>{w["risk_level"].upper()}</b></font>'
                f' &nbsp;&middot;&nbsp; '
                f'<font color="{hx(effort_color)}"><b>EFFORT {w["effort"].upper()}</b></font>'
                f' &nbsp;&middot;&nbsp; '
                f'<font color="{hx(impact_color)}"><b>IMPACT {w["impact"].upper()}</b></font>'
            )
            why_html = f'<b>Why:</b> {w["why"]}'
            action_html = f'<b>Next:</b> {w["next_action"]}'

            # One paragraph per row, single column. Avoids the previous nested-
            # column layout that overflowed when title + badges combined exceeded
            # the card's usable width.
            qw_card = Table(
                [[Paragraph(header_html, body_style)],
                 [Paragraph(badges_html, badges_style)],
                 [Paragraph(why_html, body_muted_style)],
                 [Paragraph(action_html, body_muted_style)]],
                colWidths=[6.5 * inch],
            )
            qw_card.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), BRAND_LIGHT),
                ('LEFTPADDING', (0, 0), (-1, -1), 14),
                ('RIGHTPADDING', (0, 0), (-1, -1), 14),
                ('TOPPADDING', (0, 0), (0, 0), 8),
                ('BOTTOMPADDING', (0, 0), (0, 0), 2),
                ('TOPPADDING', (0, 1), (0, 1), 0),
                ('BOTTOMPADDING', (0, 1), (0, 1), 4),
                ('TOPPADDING', (0, 2), (-1, -2), 2),
                ('BOTTOMPADDING', (0, 2), (-1, -2), 2),
                ('TOPPADDING', (0, -1), (-1, -1), 2),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
                ('LINEBEFORE', (0, 0), (0, -1), 3, INFO),
            ]))
            story.append(KeepTogether(qw_card))
            story.append(Spacer(1, 0.06 * inch))

        story.append(PageBreak())

    # ===== CONTROL DETAILS (grouped by CIS group) =====
    story.append(Paragraph("Control Details", section_style))
    story.append(Paragraph(
        "Complete inventory of all CIS Controls v8, grouped by Implementation Group, with status, risk level, and safeguard progress.",
        body_muted_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    groups_order = ["Basic", "Foundational", "Organizational"]
    by_group: dict = {g: [] for g in groups_order}
    for c in sorted(controls, key=lambda x: int(x.cis_id)):
        g = get_cis_group(c.cis_id)
        by_group.setdefault(g, []).append(c)

    cell_name_style = ParagraphStyle('CellName', parent=body_style, fontSize=9, leading=11)

    group_header_style = ParagraphStyle(
        'GroupHeader', parent=body_style, spaceAfter=4, keepWithNext=1,
    )

    for g_name in groups_order:
        members = by_group.get(g_name, [])
        if not members:
            continue

        g_implemented = sum(1 for c in members if c.status == "implemented")
        story.append(Spacer(1, 0.05 * inch))
        story.append(Paragraph(
            f'<font name="Helvetica-Bold" size="11" color="{hx(BRAND)}">{g_name}</font> '
            f'<font size="9" color="{hx(MUTED)}">'
            f'&middot; {g_implemented}/{len(members)} controls implemented</font>',
            group_header_style,
        ))

        rows = [["#", "Control", "Status", "Risk", "Safeguards", "Owner"]]
        cell_styles = []
        for ridx, c in enumerate(members, start=1):
            sg_total = len(c.safeguards)
            sg_impl = count_implemented(c.safeguards)
            owner = c.owner.full_name if c.owner else "—"
            rows.append([
                c.cis_id,
                Paragraph(c.name, cell_name_style),
                c.status.replace("_", " ").title(),
                c.risk_level.title(),
                f"{sg_impl}/{sg_total}",
                owner,
            ])
            cell_styles.append(('BACKGROUND', (2, ridx), (2, ridx), STATUS_BG.get(c.status, BRAND_LIGHT)))
            cell_styles.append(('TEXTCOLOR', (2, ridx), (2, ridx), STATUS_FG.get(c.status, MUTED)))
            cell_styles.append(('FONTNAME', (2, ridx), (2, ridx), 'Helvetica-Bold'))
            cell_styles.append(('TEXTCOLOR', (3, ridx), (3, ridx), RISK_FG.get(c.risk_level, MUTED)))
            cell_styles.append(('FONTNAME', (3, ridx), (3, ridx), 'Helvetica-Bold'))

        col_widths = [0.45 * inch, 2.7 * inch, 1.25 * inch, 0.75 * inch, 0.85 * inch, 0.95 * inch]
        ct = Table(rows, colWidths=col_widths, repeatRows=1)
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, BORDER),
            ('LINEBELOW', (0, 1), (-1, -1), 0.25, BORDER),
        ] + cell_styles))
        story.append(ct)
        story.append(Spacer(1, 0.12 * inch))

    # ----- Footers (silent on cover, branded on body)
    def cover_canvas(canvas, doc_):
        return

    def body_canvas(canvas, doc_):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(MUTED)
        text = f"{platform_name}  ·  CIS Controls v8  ·  Page {doc_.page}"
        canvas.drawCentredString(letter[0] / 2, 32, text)
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.4)
        canvas.line(54, letter[1] - 36, letter[0] - 54, letter[1] - 36)
        canvas.restoreState()

    doc.build(story, onFirstPage=cover_canvas, onLaterPages=body_canvas)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# fpdf2 fallback (kept minimal; only used when reportlab is unavailable)
# ---------------------------------------------------------------------------

def _generate_with_fpdf(db: Session) -> bytes:
    from fpdf import FPDF
    controls = db.query(Control).all()
    total = len(controls)
    implemented = sum(1 for c in controls if c.status == ControlStatus.IMPLEMENTED.value)
    score = compute_compliance_score(controls)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 12, "CIS Controls v8 Assessment Report", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", ln=True)
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Compliance Score: {score}%", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 6, f"Total: {total} | Implemented: {implemented}", ln=True)
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Control Details", ln=True)
    pdf.set_font("Helvetica", "", 10)
    for c in controls:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, f"{c.cis_id} - {c.name}", ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, f"Status: {c.status.replace('_', ' ').title()} | Risk: {c.risk_level.title()}", ln=True)
        if c.objective:
            pdf.multi_cell(0, 5, f"Objective: {c.objective}")
        pdf.ln(3)
    return pdf.output()


# ---------------------------------------------------------------------------
# XLSX (unchanged from previous version)
# ---------------------------------------------------------------------------

def generate_xlsx_report(db: Session) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    controls = db.query(Control).all()
    total = len(controls)
    implemented = sum(1 for c in controls if c.status == ControlStatus.IMPLEMENTED.value)
    in_progress = sum(1 for c in controls if c.status == ControlStatus.IN_PROGRESS.value)
    not_impl = sum(1 for c in controls if c.status == ControlStatus.NOT_IMPLEMENTED.value)
    needs_review = sum(1 for c in controls if c.status == ControlStatus.NEEDS_REVIEW.value)
    score = compute_compliance_score(controls)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E4E4E7"),
        right=Side(style="thin", color="E4E4E7"),
        top=Side(style="thin", color="E4E4E7"),
        bottom=Side(style="thin", color="E4E4E7"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    ws = wb.active
    ws.title = "Summary"
    ws.append(["CIS Controls v8 Assessment Report"])
    ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([])
    ws.append(["Compliance Score", f"{score}%"])
    ws.append(["Total Controls", total])
    ws.append(["Implemented", implemented])
    ws.append(["In Progress", in_progress])
    ws.append(["Not Implemented", not_impl])
    ws.append(["Needs Review", needs_review])
    ws.append([])
    ws.append(["Risk Distribution"])
    ws.append(["Critical", sum(1 for c in controls if c.risk_level == "critical")])
    ws.append(["High", sum(1 for c in controls if c.risk_level == "high")])
    ws.append(["Medium", sum(1 for c in controls if c.risk_level == "medium")])
    ws.append(["Low", sum(1 for c in controls if c.risk_level == "low")])
    for cell in ws[4]:
        cell.font = Font(bold=True)
    for cell in ws[11]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet(title="Controls")
    headers = ["CIS ID", "Name", "Group", "Status", "Risk", "Safeguards", "Owner", "Objective"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for c in sorted(controls, key=lambda x: int(x.cis_id)):
        sg_total = len(c.safeguards)
        sg_impl = count_implemented(c.safeguards)
        owner = c.owner.full_name if c.owner else "—"
        ws2.append([
            c.cis_id, c.name, get_cis_group(c.cis_id),
            c.status.replace("_", " ").title(), c.risk_level.title(),
            f"{sg_impl}/{sg_total}", owner, c.objective or "",
        ])

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        max_len = 0
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    ws3 = wb.create_sheet(title="Safeguards")
    headers3 = ["CIS ID", "Control Name", "Safeguard ID", "Title", "IG", "Status"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for c in sorted(controls, key=lambda x: int(x.cis_id)):
        for s in c.safeguards:
            ws3.append([
                c.cis_id, c.name, s.safeguard_id, s.title,
                s.ig.upper(),
                s.implementation_status.replace("_", " ").title(),
            ])

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    for col in range(1, len(headers3) + 1):
        max_len = 0
        for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws3.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
